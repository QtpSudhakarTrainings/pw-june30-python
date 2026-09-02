from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_excel_file(file_path: str) -> list[dict[str, Any]]:
    """Reads the first sheet of an .xlsx file and returns it as a list of row dicts, keyed by
    the header row — the Python equivalent of DataUtils/ExcelDataUtil.ts's XLSX.utils.sheet_to_json.
    """
    workbook = load_workbook(Path(file_path), read_only=True, data_only=True)
    sheet = workbook.worksheets[0]

    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)

    data: list[dict[str, Any]] = []
    for row in rows:
        if all(value is None for value in row):
            continue
        data.append(dict(zip(headers, row)))

    workbook.close()
    return data
